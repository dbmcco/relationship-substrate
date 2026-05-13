from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from relationship_substrate.contracts import SourceEventIn, SourcePosture


def _normalize_header(value: object) -> str:
    return str(value or "").lstrip("\ufeff").strip().lower().replace(" ", "_").replace("-", "_")


def _value(record: dict[str, object], *keys: str) -> object:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return None


def _event_from_record(
    *,
    record: dict[str, object],
    path: Path,
    container: str,
    row_number: int,
) -> SourceEventIn | None:
    email = _value(record, "email", "work_email", "email_address", "e_mail")
    first_name = _value(record, "first_name")
    last_name = _value(record, "last_name")
    full_name = _value(record, "name", "full_name")
    company = _value(record, "company", "organization")
    title = _value(record, "title", "job_title")
    if not email and not first_name and not last_name and not full_name and not company:
        return None
    payload = {
        "first_name": first_name,
        "last_name": last_name,
        "full_name": full_name,
        "title": title,
        "company": company,
        "email": email,
        "row_number": row_number,
        "path": str(path),
        "source_container": container,
    }
    return SourceEventIn(
        source_name="next_up",
        source_event_type="curated_contact",
        source_event_key=f"next_up:{path.name}:{container}:{row_number}",
        source_payload=payload,
        source_posture=SourcePosture.CURATED_EXPORT,
        provenance_status="unknown_upstream",
        trust_role="identity/context seed",
    )


def iter_people_workbook_events(path: Path) -> Iterator[SourceEventIn]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Contacts"] if "Contacts" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [_normalize_header(value) for value in next(rows)]
    for row_number, values in enumerate(rows, start=2):
        record = dict(zip(headers, values, strict=False))
        event = _event_from_record(
            record=record,
            path=path,
            container=sheet.title,
            row_number=row_number,
        )
        if event is not None:
            yield event


def iter_people_csv_events(path: Path) -> Iterator[SourceEventIn]:
    with path.open(newline="", errors="replace") as handle:
        reader = csv.DictReader(handle)
        for row_number, record in enumerate(reader, start=2):
            normalized = {_normalize_header(key): value for key, value in record.items()}
            event = _event_from_record(
                record=normalized,
                path=path,
                container=path.stem,
                row_number=row_number,
            )
            if event is not None:
                yield event


def iter_next_up_events(path: Path) -> Iterator[SourceEventIn]:
    paths: list[Path]
    if path.is_dir():
        paths = sorted(
            [
                *path.glob("*.xlsx"),
                *path.glob("*.csv"),
            ]
        )
    else:
        paths = [path]
    for item in paths:
        if item.suffix.lower() == ".xlsx":
            yield from iter_people_workbook_events(item)
        elif item.suffix.lower() == ".csv":
            yield from iter_people_csv_events(item)
