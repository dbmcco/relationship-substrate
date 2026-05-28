from __future__ import annotations

import json
import sys
from pathlib import Path
from uuid import uuid4

import psycopg
from openpyxl import Workbook

from relationship_substrate.cli import main
from relationship_substrate.db import run_migrations


def _workbook(path: Path, email: str = "Jane@Example.com") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["First Name", "Last Name", "Title", "Company", "Email"])
    ws.append(["Jane", "Doe", "VP Product", "ExampleCo", email])
    wb.save(path)
    return path


def _run_cli(monkeypatch, capsys, *args: str) -> dict:
    monkeypatch.setattr(sys, "argv", ["relationship-substrate", *args])
    assert main() == 0
    return json.loads(capsys.readouterr().out)


def test_agent_cli_ingests_materializes_and_exports_from_db(
    database_url, tmp_path, monkeypatch, capsys
):
    run_migrations(database_url)
    workbook = _workbook(tmp_path / "people.xlsx")
    event_key = f"next_up:{workbook.name}:Contacts:2"
    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM relationship_substrate.source_event WHERE source_event_key = %s",
                (event_key,),
            )
            cur.execute(
                "DELETE FROM relationship_substrate.person WHERE primary_email = %s",
                ("jane@example.com",),
            )
        conn.commit()

    ingest = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-next-up",
        "--path",
        str(workbook),
    )
    assert ingest == {"source": "next_up", "events_seen": 1, "events_upserted": 1}

    materialized = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-exact-emails",
        "--source",
        "next_up",
    )
    assert materialized["source"] == "next_up"
    assert materialized["materialized"] >= 1

    picture = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "export-operating-picture",
        "--from-db",
        "--limit",
        "100000",
    )
    assert picture["relationships"]
    relationship = next(
        row for row in picture["relationships"] if row["metadata"]["primary_email"] == "jane@example.com"
    )
    assert relationship["name"] == "Jane Doe"
    assert relationship["relationship_state"] == "uninterpreted_identity_seed"


def test_agent_cli_eval_local_writes_machine_readable_artifacts(
    database_url, tmp_path, monkeypatch, capsys
):
    run_migrations(database_url)
    workbook = _workbook(tmp_path / "eval_people.xlsx", email="eval@example.com")
    output_dir = tmp_path / "eval-output"

    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "eval-local",
        "--next-up-path",
        str(workbook),
        "--output-dir",
        str(output_dir),
        "--skip-msgvault",
    )

    assert report["ok"] is True
    assert report["next_up"]["events_seen"] == 1
    assert report["materialization"]["materialized"] >= 1
    assert report["operating_picture"]["relationships"] >= 1
    assert report["identity_candidates"]["source"] == "identity_candidate"
    assert (output_dir / "eval_report.json").exists()
    assert (output_dir / "relationship_operating_picture.json").exists()


def test_agent_cli_eval_local_can_include_calendar_json(database_url, tmp_path, monkeypatch, capsys):
    run_migrations(database_url)
    localpart = f"eval-calendar-{uuid4().hex}"
    email = f"{localpart}@example.com"
    workbook = _workbook(tmp_path / "eval_calendar_people.xlsx", email=email)
    calendar_path = tmp_path / "eval-calendar.json"
    calendar_path.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "id": f"event-{localpart}",
                        "summary": "Eval calendar event",
                        "start": {"dateTime": "2026-05-03T10:00:00-04:00"},
                        "attendees": [
                            {"email": "user@example.com", "self": True},
                            {"email": email, "displayName": "Eval Calendar"},
                        ],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    output_dir = tmp_path / "eval-calendar-output"

    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "eval-local",
        "--next-up-path",
        str(workbook),
        "--calendar-path",
        str(calendar_path),
        "--output-dir",
        str(output_dir),
        "--skip-msgvault",
    )

    assert report["calendar"]["ingestion"]["events_seen"] == 1
    assert report["calendar"]["materialization"]["attendees_materialized"] >= 1


def test_agent_cli_generates_identity_candidates(database_url, monkeypatch, capsys):
    run_migrations(database_url)
    localpart = f"cliidentitycandidate{uuid4().hex}"
    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO relationship_substrate.person (
                  display_name, primary_email, source_posture, provenance_status
                )
                VALUES
                  ('CLI Identity Candidate', %s, 'direct_interaction', 'test'),
                  ('CLI Identity Candidate', %s, 'direct_interaction', 'test')
                ON CONFLICT (primary_email)
                DO UPDATE SET display_name = EXCLUDED.display_name
                """,
                (f"{localpart}@example.com", f"{localpart}@other.example"),
            )
        conn.commit()

    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "generate-identity-candidates",
    )

    assert report["source"] == "identity_candidate"
    assert report["candidate_pairs"] >= 1
    assert report["open_candidates"] >= 1


def test_agent_cli_reviews_identity_candidate(database_url, monkeypatch, capsys):
    run_migrations(database_url)
    localpart = f"clireviewcandidate{uuid4().hex}"
    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO relationship_substrate.person (
                  display_name, primary_email, source_posture, provenance_status
                )
                VALUES
                  ('CLI Review Candidate', %s, 'direct_interaction', 'test'),
                  ('CLI Review Candidate', %s, 'direct_interaction', 'test')
                ON CONFLICT (primary_email)
                DO UPDATE SET display_name = EXCLUDED.display_name
                """,
                (f"{localpart}@example.com", f"{localpart}@other.example"),
            )
        conn.commit()

    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "generate-identity-candidates",
    )
    listed = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "list-identity-candidates",
        "--limit",
        "1000",
    )
    candidate = next(row for row in listed["candidates"] if row["evidence"]["match_key"] == localpart)

    shown = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "show-identity-candidate",
        "--id",
        candidate["id"],
    )
    resolved = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "resolve-identity-candidate",
        "--id",
        candidate["id"],
        "--status",
        "rejected",
        "--note",
        "Not enough evidence to merge.",
    )

    assert shown["id"] == candidate["id"]
    assert resolved["status"] == "rejected"
    assert resolved["evidence"]["review"]["note"] == "Not enough evidence to merge."


def test_agent_cli_ingests_and_materializes_calendar_json(database_url, tmp_path, monkeypatch, capsys):
    run_migrations(database_url)
    calendar_path = tmp_path / "calendar.json"
    localpart = f"calcli{uuid4().hex}"
    calendar_path.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "id": f"event-{localpart}",
                        "summary": "Calendar CLI smoke",
                        "start": {"dateTime": "2026-05-02T10:00:00-04:00"},
                        "attendees": [
                            {"email": "user@example.com", "self": True},
                            {"email": f"{localpart}@example.com", "displayName": "Calendar Person"},
                        ],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    ingested = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-calendar",
        "--path",
        str(calendar_path),
    )
    materialized = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-calendar-events",
    )
    picture = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "export-operating-picture",
        "--from-db",
        "--limit",
        "1000",
    )

    assert ingested == {"source": "calendar", "events_seen": 1, "events_upserted": 1}
    assert materialized["attendees_materialized"] >= 1
    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT count(*)
                FROM relationship_substrate.evidence_ref
                WHERE ref_type = 'calendar_event'
                AND ref_value = %s
                """,
                (f"calendar:{calendar_path.name}:event-{localpart}",),
            )
            assert cur.fetchone() == (1,)
    relationship = next(
        row for row in picture["relationships"] if row["metadata"]["primary_email"] == f"{localpart}@example.com"
    )
    assert relationship["metadata"]["calendar_interaction_count"] == 1


def test_agent_cli_ingests_calendar_page_directory(database_url, tmp_path, monkeypatch, capsys):
    run_migrations(database_url)
    export_dir = tmp_path / "calendar-pages"
    export_dir.mkdir()
    first = f"clipageone{uuid4().hex}"
    second = f"clipagetwo{uuid4().hex}"
    for name, localpart in [("page-1.json", first), ("page-2.json", second)]:
        (export_dir / name).write_text(
            json.dumps(
                {
                    "items": [
                        {
                            "id": f"event-{localpart}",
                            "summary": "Calendar page smoke",
                            "start": {"dateTime": "2026-05-02T10:00:00-04:00"},
                            "attendees": [
                                {"email": f"{localpart}@example.com", "displayName": "Calendar Page Person"},
                            ],
                        }
                    ]
                }
            ),
            encoding="utf-8",
        )

    ingested = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-calendar",
        "--path",
        str(export_dir),
    )
    materialized = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-calendar-events",
    )

    assert ingested == {
        "source": "calendar",
        "files_seen": 2,
        "events_seen": 2,
        "events_upserted": 2,
    }
    assert materialized["attendees_materialized"] >= 2


def test_agent_cli_shows_person_dossier(database_url, tmp_path, monkeypatch, capsys):
    run_migrations(database_url)
    calendar_path = tmp_path / "dossier-calendar.json"
    localpart = f"clidossier{uuid4().hex}"
    email = f"{localpart}@example.com"
    calendar_path.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "id": f"event-{localpart}",
                        "summary": "CLI dossier meeting",
                        "start": {"dateTime": "2026-05-06T11:00:00-04:00"},
                        "attendees": [
                            {"email": "user@example.com", "self": True},
                            {"email": email, "displayName": "CLI Dossier"},
                        ],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-calendar",
        "--path",
        str(calendar_path),
    )
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-calendar-events",
    )

    dossier = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "show-person",
        "--email",
        email,
    )

    assert dossier["person"]["primary_email"] == email
    assert dossier["relationship_edge"]["calendar_interaction_count"] == 1
    assert dossier["interactions"][0]["subject"] == "CLI dossier meeting"


def test_agent_cli_records_and_lists_person_notes(database_url, monkeypatch, capsys):
    run_migrations(database_url)
    email = f"cli-person-note-{uuid4().hex}@example.com"
    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO relationship_substrate.person (
                  display_name, primary_email, source_posture, provenance_status
                )
                VALUES ('CLI Person Note', %s, 'direct_interaction', 'test')
                """,
                (email,),
            )
        conn.commit()

    recorded = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "record-person-note",
        "--person",
        email,
        "--kind",
        "context_fit",
        "--applies-to",
        "small_consulting_firm_discovery",
        "--note",
        "This is known context the agent should carry forward.",
    )
    listed = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "list-person-notes",
        "--person",
        email,
    )

    assert recorded["person_email"] == email
    assert listed["count"] == 1
    assert listed["notes"][0]["note_kind"] == "context_fit"


def test_agent_cli_records_and_lists_subject_notes(database_url, monkeypatch, capsys):
    run_migrations(database_url)
    email = f"cli-subject-note-{uuid4().hex}@example.com"
    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO relationship_substrate.person (
                  display_name, primary_email, source_posture, provenance_status
                )
                VALUES ('CLI Subject Note', %s, 'direct_interaction', 'test')
                """,
                (email,),
            )
        conn.commit()

    recorded = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "record-subject-note",
        "--subject-type",
        "person",
        "--subject",
        email,
        "--kind",
        "context_fit",
        "--applies-to",
        "small_consulting_firm_discovery",
        "--note",
        "This source-owned correction should be contextual evidence.",
    )
    listed = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "list-subject-notes",
        "--subject-type",
        "person",
        "--subject",
        email,
    )

    assert recorded["person_email"] == email
    assert listed["count"] == 1
    assert listed["subject_note_context"][0]["note_kind"] == "context_fit"
    assert listed["notes"] == listed["subject_note_context"]


def test_agent_cli_searches_people_by_role_and_company_size(
    database_url, tmp_path, monkeypatch, capsys
):
    run_migrations(database_url)
    localpart = uuid4().hex
    workbook = tmp_path / "search_people.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["First Name", "Last Name", "Title", "Company", "Email"])
    for index in range(10):
        title = "Principal Consultant" if index == 0 else "Delivery Lead"
        ws.append(
            [
                f"Search{index}",
                "Person",
                title,
                f"CLI Search Advisors {localpart}",
                f"cli-search-{index}-{localpart}@example.com",
            ]
        )
    wb.save(workbook)

    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-next-up",
        "--path",
        str(workbook),
    )
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-exact-emails",
    )

    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "search-people",
        "--role-keywords",
        "consultant",
        "--known-people-at-company-min",
        "10",
        "--known-people-at-company-max",
        "15",
        "--limit",
        "1000",
    )

    result = next(
        row for row in report["results"] if row["email"] == f"cli-search-0-{localpart}@example.com"
    )
    assert result["known_people_at_company_count"] == 10


def test_agent_cli_embeds_curated_contacts_with_hash_provider(
    database_url, tmp_path, monkeypatch, capsys
):
    run_migrations(database_url)
    workbook = _workbook(
        tmp_path / "embed_people.xlsx",
        email=f"cli-embed-{uuid4().hex}@example.com",
    )
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-next-up",
        "--path",
        str(workbook),
    )
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-exact-emails",
    )

    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "embed-curated-contacts",
        "--provider",
        "hash",
    )

    assert report["source"] == "next_up"
    assert report["embedded"] >= 1
    assert report["provider"] == "hash"


def test_agent_cli_proposes_relationship_state_live_with_registry_route(
    database_url, monkeypatch, capsys
):
    run_migrations(database_url)
    captured: dict[str, object] = {}

    def fake_live_proposal(
        db_url: str,
        *,
        email: str,
        route_key: str,
        service_name: str,
        evidence_limit: int,
        registry_path: str | None = None,
    ) -> dict[str, object]:
        captured["db_url"] = db_url
        captured["email"] = email
        captured["route_key"] = route_key
        captured["service_name"] = service_name
        captured["evidence_limit"] = evidence_limit
        captured["registry_path"] = registry_path
        return {
            "model_route": {"route_key": route_key},
            "proposal_event": {"source_event_id": "proposal-id"},
            "relationship_state": {"id": "state-id"},
            "journal_entry": {"id": "journal-id"},
        }

    monkeypatch.setattr("relationship_substrate.cli.propose_relationship_state_live", fake_live_proposal)
    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "propose-relationship-state-live",
        "--email",
        "person@example.com",
        "--model-route",
        "relationship_substrate.relationship_state_proposal",
        "--registry-service",
        "relationship-substrate",
        "--registry-path",
        "/tmp/cognition-presets.toml",
        "--evidence-limit",
        "8",
    )

    assert report["relationship_state"]["id"] == "state-id"
    assert captured == {
        "db_url": database_url,
        "email": "person@example.com",
        "route_key": "relationship_substrate.relationship_state_proposal",
        "service_name": "relationship-substrate",
        "evidence_limit": 8,
        "registry_path": "/tmp/cognition-presets.toml",
    }


def test_agent_cli_upserts_organization_enrichment(
    database_url, tmp_path, monkeypatch, capsys
):
    run_migrations(database_url)
    localpart = uuid4().hex
    company = f"CLI Enterprise Pharma {localpart}"
    workbook = tmp_path / "org_enrichment_people.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["First Name", "Last Name", "Title", "Company", "Email"])
    for index in range(10):
        title = "Medical Communications Lead" if index == 0 else "Peer"
        ws.append(
            [
                f"Org{index}",
                "Person",
                title,
                company,
                f"cli-org-{index}-{localpart}@example.com",
            ]
        )
    wb.save(workbook)
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-next-up",
        "--path",
        str(workbook),
    )
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-exact-emails",
    )
    enrichment = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "upsert-organization-enrichment",
        "--company",
        company,
        "--company-type",
        "public_pharmaceutical_company",
        "--employee-count-min",
        "50000",
        "--employee-count-label",
        "enterprise",
        "--consultant-count-estimate",
        "14",
        "--source-name",
        "manual_research",
        "--source-url",
        "https://example.com/company-profile",
    )
    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "search-people",
        "--role-keywords",
        "medical communications",
        "--known-people-at-company-min",
        "10",
        "--known-people-at-company-max",
        "15",
        "--limit",
        "1000",
    )

    result = next(
        row for row in report["results"] if row["email"] == f"cli-org-0-{localpart}@example.com"
    )
    assert enrichment["enrichment"]["company_type"] == "public_pharmaceutical_company"
    assert result["known_people_at_company_count"] == 10
    assert result["organization_enrichment"]["employee_count_label"] == "enterprise"
    assert result["organization_enrichment"]["consultant_count_estimate"] == 14

    consultant_report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "search-people",
        "--role-keywords",
        "medical communications",
        "--consultant-count-min",
        "10",
        "--consultant-count-max",
        "20",
        "--limit",
        "1000",
    )

    consultant_result = next(
        row
        for row in consultant_report["results"]
        if row["email"] == f"cli-org-0-{localpart}@example.com"
    )
    assert "consultant_count_estimate:14" in consultant_result["match_reasons"]


def test_agent_cli_exports_and_imports_organization_enrichment_batch(
    database_url, tmp_path, monkeypatch, capsys
):
    run_migrations(database_url)
    localpart = uuid4().hex
    company = f"Batch Medcom Co {localpart}"
    workbook = tmp_path / "batch_org_people.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["First Name", "Last Name", "Title", "Company", "Email"])
    ws.append(["Batch", "Person", "Medical Communications Consultant", company, f"batch-{localpart}@example.com"])
    wb.save(workbook)
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-next-up",
        "--path",
        str(workbook),
    )

    worklist = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "export-organization-enrichment-worklist",
        "--limit",
        "1000",
    )
    row = next(item for item in worklist["companies"] if item["company_name"] == company)
    assert row["has_enrichment"] is False
    assert row["sample_titles"] == ["Medical Communications Consultant"]

    import_path = tmp_path / "org-enrichment.json"
    import_path.write_text(
        json.dumps(
            [
                {
                    "company_name": company,
                    "company_type": "medical_communications_consultancy",
                    "employee_count_min": 10,
                    "employee_count_max": 20,
                    "employee_count_label": "small_team",
                    "source_name": "manual_research",
                    "source_url": "https://example.com/company",
                    "provenance_status": "external_research",
                }
            ]
        ),
        encoding="utf-8",
    )
    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "import-organization-enrichments",
        "--path",
        str(import_path),
    )

    assert report["imported"] == 1
    assert report["skipped"] == 0


def test_agent_cli_exports_history_backed_organization_worklist(
    database_url, tmp_path, monkeypatch, capsys
):
    run_migrations(database_url)
    localpart = uuid4().hex
    company = f"CLI History Co {localpart}"
    domain = f"historyco-{localpart}.example"
    email = f"history-{localpart}@{domain}"
    workbook = tmp_path / "history_org_people.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["First Name", "Last Name", "Title", "Company", "Email"])
    ws.append(["History", "Person", "Medical Communications Consultant", company, email])
    wb.save(workbook)
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "ingest-next-up",
        "--path",
        str(workbook),
    )
    _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "materialize-exact-emails",
    )
    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO relationship_substrate.relationship_edge (
                  person_id, interaction_count, metadata
                )
                SELECT id, 7, '{"source": "test"}'::jsonb
                FROM relationship_substrate.person
                WHERE primary_email = %s
                ON CONFLICT (person_id)
                DO UPDATE SET interaction_count = EXCLUDED.interaction_count
                """,
                (email,),
            )
        conn.commit()

    report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "export-history-backed-organization-worklist",
        "--limit",
        "1000",
    )

    row = next(item for item in report["companies"] if item["company_name"] == company)
    assert row["domain"] == domain
    assert row["known_people_count"] == 1
    assert row["direct_people_count"] == 1
    assert row["total_interaction_count"] == 7
    assert row["strongest_people"][0]["email"] == email

    missing_only_report = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "export-history-backed-organization-worklist",
        "--missing-only",
        "--limit",
        "1000",
    )
    assert any(item["company_name"] == company for item in missing_only_report["companies"])


def test_agent_cli_prepares_relationship_tone_tenor_analysis_packet(database_url, monkeypatch, capsys):
    run_migrations(database_url)
    email = f"tone-cli-{uuid4().hex}@example.com"

    with psycopg.connect(database_url) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO relationship_substrate.person (
                  display_name, primary_email, source_posture, provenance_status
                )
                VALUES (%s, %s, 'direct_interaction', 'test')
                ON CONFLICT (primary_email)
                DO UPDATE SET display_name = EXCLUDED.display_name
                """,
                ("Tone CLI Person", email),
            )
        conn.commit()

    packet = _run_cli(
        monkeypatch,
        capsys,
        "--database-url",
        database_url,
        "prepare-relationship-tone-analysis",
        "--email",
        email,
        "--evidence-limit",
        "3",
        "--prior-state-limit",
        "2",
    )

    assert packet["analysis_stage"] == "relationship_tone_tenor"
    assert packet["count"] == 1
    assert packet["people"][0]["email"] == email
